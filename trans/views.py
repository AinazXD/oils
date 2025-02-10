# Справочник данных по насосам типа НМ
from django.http import HttpResponseNotAllowed
import json
from .models import MainPump, BoosterPump
from .formulas import (
    calculate_density_at_temperature,
    calculate_u,
    calculate_viscosity_at_temperature,
    calculate_hourly_rate,
    flow_rate,
    calculate_reynolds_number,
    calculate_relative_roughness,
    calculate_transition_reynolds_numbers,
    calculate_hydraulic_resistance,
    calculate_velocity,
    calculate_friction_head_loss,
    calculate_total_head_loss,

)
from django.contrib.auth.decorators import login_required
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse

from .models import Pipeline, Product
from django.contrib.auth.decorators import login_required

from django.shortcuts import render
from django.contrib.auth import login, authenticate
from .forms import RegisterForm, LoginForm
from django.contrib.auth import logout
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required

import matplotlib
matplotlib.use('Agg')  # Установка бэкенда Agg
import matplotlib.pyplot as plt
from io import BytesIO
import base64

NM_PUMP_DATA = {
    "НМ 1250-260": [
        {"Rotor": 1, "H_n": 318.8, "a": 38.7, "b": 0.017, "a₀": 1.037, "b₀": 20.29, "c₀": 10.36, "c₁": -44.35, "c₂": 26, "Dвх": 353, "D₂": 440, "nₑ": 73},
        {"Rotor": 0.7, "H_n": 283, "a": 35.4, "b": 0.163, "a₀": 0.678, "b₀": 17.14, "c₀": 11.91, "c₁": -52.68, "c₂": 26, "Dвх": 353, "D₂": 418, "nₑ": 59},
        {"Rotor": 0.7, "H_n": 216.4, "a": 40.9, "b": 0.092, "a₀": 0.76, "b₀": 9.63, "c₀": 14.3, "c₁": -69.6, "c₂": None, "Dвх": 353, "D₂": 418, "nₑ": 62},
        {"Rotor": 1, "H_n": 316.8, "a": 41.9, "b": 0.092, "a₀": 0.76, "b₀": 20.6, "c₀": 11.3, "c₁": -50, "c₂": None, "Dвх": 353, "D₂": 460, "nₑ": 71},
        {"Rotor": 1, "H_n": 289.8, "a": 34.8, "b": 0.092, "a₀": 0.76, "b₀": 20.6, "c₀": 11.3, "c₁": -50, "c₂": None, "Dвх": 353, "D₂": 418, "nₑ": 77},
        {"Rotor": 1, "H_n": 271, "a": 43.9, "b": 0.092, "a₀": 0.76, "b₀": 20.6, "c₀": 11.3, "c₁": -50, "c₂": None, "Dвх": 353, "D₂": 395, "nₑ": 89},
        {"Rotor": 1.25, "H_n": 327.4, "a": 25.0, "b": 1.17, "a₀": 0.46, "b₀": 34.10, "c₀": None, "c₁": -21.7, "c₂": None, "Dвх": 353, "D₂": 450, "nₑ": 79},
    ],
    "НМ 1250-400": [
        {"Rotor": 1, "H_n": 544.3, "a": 9.24, "b": 6.071, "a₀": 0.245, "b₀": -1.32, "c₀": 15.2, "c₁": -67.56, "c₂": None, "Dвх": 353, "D₂": 365, "nₑ": 41},
    ],
    "НМ 1800-240": [
        {"Rotor": 1, "H_n": 298.7, "a": 17.8, "b": 1.29e-4, "a₀": 1.634, "b₀": 3.86, "c₀": 9.51, "c₁": -28.57, "c₂": None, "Dвх": 512, "D₂": 440, "nₑ": 90},
    ],
    "НМ 2500-230": [
        {"Rotor": 1, "H_n": 287.9, "a": 9.47, "b": 0.0598, "a₀": 0.813," b₀" :6.86," c₀" :7.11," c₁" :-15.63," c₂" :36.2," Dвх" :512," D₂" :430," nₑ" :109},
        {"Rotor" :0.7," H_n" :246.7," a" :7.18," b" :0.487," a₀" :0.538," b₀" :4.96," c₀" :7.944," c₁" :-19.81," c₂" :38," Dвх" :512," D₂" :405," nₑ" :95},
        {"Rotor" :0.5," H_n" :248," a" :16.3," b" :1.273," a₀" :0.399," b₀" :5.66," c₀" :9.73," c₁" :-29.87," c₂" :26," Dвх" :512," D₂" :425," nₑ" :79},
        {"Rotor" :0.5," H_n" :246.7," a" :16.8," b" :1.41," a₀" :0.39," b₀" :24.8," c₀" :644," c₁" :-16.9," c₂" :None," Dвх" :512," D₂" :425," nₑ" :77},
        {"Rotor" :0.7," H_n" :248.7," a" :7.61," b" :1.97," a₀" :0.35," b₀" :-79.35," c₀" :15.8," c₁" :-37," c₂" :None," Dвх" :512," D₂" :405," nₑ" :93},
        {"Rotor" :1," H_n" :281.5," a" :7.84," b" :1.26," a₀" :0.42," b₀" :26.2," c₀" :485," c₁" :-9.7," c₂" :None," Dвх" :512," D₂" :440," nₑ" :109},
        {"Rotor" :None," H_n" :258.8," a" :8.59," b" :1.26," a₀" :0.42," b₀" :26.2," c₀" :485," c₁" :-9.7," c₂" :None," Dвх" :512," D₂" :"440"," nₑ":"117"},
        {"Rotor":"None"," H_n":"235"," a":"8"," b":"32"," a₀":"18"," b₀":"485"," c₀":"-9"," c₁":"-512"," c₂":"385"," Dвх":"450"," D₂":"129"},
        {"Rotor":"1"," H_n":"371"," a":"14"," b":"26"," a₀":"18"," b₀":"403"," c₀":"-6"," c₁":"-512"," c₂":"450"," Dвх":"123"},
    ],
    'НМ 3600-230': [
        {"Rotor":"1"," H_n":"325"," a":"7"," b":"761"," a₀":"492"," b₀":"705"," c₀":"53"," c₁":"-864"," c₂":"41"," Dвх":"512"," D₂":"450"," nₑ":"127"},
        {"Rotor":"0"," H_n":"269"," a":"724"," b":"1872"," a₀":"376"," b₀":"429"," c₀":"632"," c₁":"1228"," c₂":"43"," Dвх":"512"," D₂":"450"," nₑ":"104"},
        {"Rotor":"5"," H_n":"272"," a":"131"," b":"388"," a₀":"30"," b₀":"755"," c₀":"762"," c₁":"1952"," c₂":"29"," Dвх":"512"," D₂":"450"," nₑ":"93"},
    ],
}



# Справочник данных по насосам типа НПВ
NPV_PUMP_DATA = { "типоразмер насоса": [
        "НПВ 150-60", "НПВ 300-60", "НПВ 600-60", "НПВ 1250-60*", "НПВ 1250-60",
        "НПВ 2500-80*", "НПВ 2500-80", "НПВ 3600-90*", "НПВ 3600-90", "НПВ 5000-120*", "НПВ 5000-120"
    ],
    "Н_0,м": [
        [78.5, 63.9], [78.5, 63.1], [75.3, 62.1], [77.1, 64.2], [74.8, 69.2, 59.9],
        [113.3, 82.9], [79.7, 96.4, 86.3], [136.3, 101.8], [127, 111, 93.7],
        [151.9, 121.1], [151.3, 132.7, 120.7]
    ],
    "а,ч/м^2": [
        ["-", "-"], ["-", "-"], ["-", "-"], ["-", "—"], ["-", "-", "-"],
        ["-", "—"], ["-", "-", "-"], ["-", "—"], ["-", "-", "-"],
        ["-", "—"], ["-", "-", "-"]
    ],
    "10^6*b ,ч/м^2": [
        [836, 875], [199, 197], [45, 47.7], [11.48, 13.27], [9.5, 10.6, 8.9],
        [5.36, 3.61], [1.0, 4.5, 4.4], [3.70, 3.0], [2.9, 2.6, 1.4],
        [1.3, 1.24], [1.3, 0.09, 1.0]
    ],
    "а0,м": [
        [3, 3], [4, 4], [4, 4], [2.2, 2.2], [2.3, 2.3, 2.3],
        [3.2, 3.2], [3.3, 3.3, 3.3], [4.8, 4.8], [4.9, 4.9, 4.9],
        [5.0, 5.0], [5, 5, 5]
    ],
    "b_0": [
        ["—", "—"], ["—", "—"], ["—", "—"], ["-", "-"], ["-", "-", "-"],
        ["-", "—"], ["-", "-", "-"], ["-", "-"], ["-", "-", "-"],
        ["-", "—"], ["-", "-", "-"]
    ],
    "10^2*C_0": [
        [-4.9, 7.07], [1.0, -15.7], [9.15, 9.03], [5.0, 5.0], [17.2, 17.2, 17.2],
        [-0.75, -0.75], [32.3, 32.3, 32.3], [1.02, 1.02], [-3.64, -3.64, -3.64],
        [3.71, 3.71], [22.4, 22.4, 22.4]
    ],
    "10^4*c1": [
        [0.99, 0.73], [0.47, 0.57], [0.24, 0.20], [10.01, 10.01], [0.08, 0.08, 0.08],
        [6.93, 6.93], [0.04, 0.04, 0.04], [4.79, 4.79], [0.045, 0.045, 0.045],
        [3.54, 3.54], [0.026, 0.026, 0.026]
    ],
    "10^8*C2": [
        [-31.5, -23.7], [-7.51, -9.6], [-2.09, -1.62], [-35.11, -35.11], [-0.24, -0.24, -0.24],
        [-14.4, -14.4], [-0.081, -0.081, -0.081], [-6.69, -6.69], [-0.064, -0.064, -0.064],
        [-3.81, -3.81], [-0.027, -0.027, -0.027]
    ],
    "b_2": [
        ["", ""], ["", ""], ["", ""], ["", ""], ["", "", ""],
        ["", ""], ["", "", ""], ["", ""], ["", "", ""],
        ["", ""], ["", "", ""]
    ],
    "Д_вх": [
        [307, 307], [307, 307], [408, 408], [800, 800], [408, 408, 800],
        [800, 800], [800, 800, 800], [1000, 1000], [1000, 1000, 1000],
        [1000, 1000], [1000, 1000, 1000]
    ],
    "Д_2": [
        [230, 207], [240, 216], [445, 400], [495, 445], [525, 500, 475],
        [530, 477], [540, 515, 487], [640, 550], [610, 580, 550],
        [640, 576], [645, 613, 580]
    ],
    "n_g": [
        [103, 130], [145, 190], [103, 127], [106, 136], [106, 116, 127],
        [121, 150], [121, 131, 150], [133, 173], [133, 149, 169],
        [126, 156], [126, 139, 154]
    ]
}


def result(request):
    # Инициализация контекста
    context = {}

    # Получаем данные из сессии
    pipeline_data = request.session.get('pipeline_data', {})
    product_data = request.session.get('product_data', [])
    total_quantity = request.session.get('total_quantity', 0)

    # Проверяем, есть ли диаметр в pipeline_data
    if 'diameter' not in pipeline_data:
        return redirect('trans:calculator')

    # Получаем диаметр из данных о трубопроводе (в мм и переводим в метры)
    diameter_mm = pipeline_data['diameter']  # Диаметр в миллиметрах
    diameter_m = diameter_mm / 1000  # Переводим диаметр в метры

    # Получаем длину трубопровода (в км и переводим в метры)
    length_km = pipeline_data['length']  # Длина в километрах
    length_m = length_km * 1000  # Переводим длину в метры

    # Получаем температуру из данных о трубопроводе
    temperature = pipeline_data.get('temperature', 293)

    # Получаем перепад высот (elevation_difference) и остаточный напор (residual_head)
    elevation_difference = float(pipeline_data.get('elevation_difference', 0))  # Перепад высот (м)
    residual_head = float(pipeline_data['residual_head'])  # Остаточный напор (м)

    # Обновляем данные о продуктах с учетом новой формулы
    updated_product_data = []
    for product in product_data:
        try:
            # Рассчитываем плотность при заданной температуре
            density_at_temperature = calculate_density_at_temperature(
                product['density293'], temperature
            )
            product['density_at_temperature'] = density_at_temperature

            # Рассчитываем коэффициент крутизны вискограммы (u)
            T1 = 293  # Температура 1 (293 К)
            T2 = 273  # Температура 2 (273 К)
            v1 = product['viscosity293']  # Вязкость при 293 К
            v2 = product['viscosity273']  # Вязкость при 273 К
            u = calculate_u(T1, T2, v1, v2)
            product['u'] = round(u, 6)  # Округляем до 6 знаков после запятой

            # Рассчитываем вязкость при заданной температуре
            v_e = product['viscosity273']  # Вязкость при 273 К
            v = calculate_viscosity_at_temperature(v_e, u, temperature)
            product['viscosity_at_temperature'] = round(v, 6)  # Округляем до 6 знаков после запятой
        except ValueError as e:
            # Если расчет невозможен, добавляем сообщение об ошибке
            product['density_at_temperature'] = None
            product['u'] = None
            product['viscosity_at_temperature'] = None
            product['error'] = str(e)
        updated_product_data.append(product)

    # Рассчитываем часовой расход (О_{час})
    hourly_rate = calculate_hourly_rate(updated_product_data)

    # Рассчитываем расход \( Q \) в \( \text{м}^3/\text{с} \)
    q_flow_rate = flow_rate(hourly_rate)

    # Рассчитываем среднюю скорость потока \( v \) и округляем до 2 знаков после запятой
    v = calculate_velocity(q_flow_rate, diameter_m)
    v_rounded = round(v, 2)  # Округляем скорость до 2 знаков после запятой

    # Находим наиболее вязкий продукт
    most_viscous_product = max(
        updated_product_data,
        key=lambda x: x['viscosity_at_temperature'] if x['viscosity_at_temperature'] is not None else 0
    )

    # Получаем вязкость наиболее вязкого продукта
    viscosity_at_temperature = most_viscous_product['viscosity_at_temperature']

    # Рассчитываем число Рейнольдса \( Re \)
    if viscosity_at_temperature is not None:
        Re = calculate_reynolds_number(v, diameter_m, viscosity_at_temperature * 1e-6)  # Переводим вязкость в м²/с
    else:
        Re = None  # Если вязкость не определена

    # Рассчитываем относительную шероховатость
    delta = 0.2 / 1000  # Абсолютная шероховатость (переводим мм в метры)
    epsilon = calculate_relative_roughness(delta, diameter_m)

    # Рассчитываем переходные числа Рейнольдса
    Re_i, Re_ii = calculate_transition_reynolds_numbers(epsilon)

    # Рассчитываем коэффициент гидравлического сопротивления и округляем до 6 знаков после запятой
    lambda_value = calculate_hydraulic_resistance(Re, epsilon)
    lambda_rounded = round(lambda_value, 6)  # Округляем λ до 6 знаков после запятой




    # Рассчитываем потерю напора на трение \( h_r \)
    friction_head_loss = calculate_friction_head_loss(lambda_value, length_m, diameter_m, v_rounded)

    # Рассчитываем полные потери напора по формуле (5.28)
    total_head_loss = calculate_total_head_loss(
        friction_head_loss,  # Потери напора на трение
        elevation_difference,  # Перепад высот (из базы данных)
        residual_head,  # Остаточный напор (из базы данных)
        N=2  # Количество использований напора
    )

    # Подбор основного насоса
    required_flow_rate = hourly_rate  # Требуемый расход (м³/ч)
    required_head = total_head_loss  # Требуемый напор (м)
    # Ищем подходящий основной насос
    main_pumps = MainPump.objects.filter(
        flow_rate__gte=required_flow_rate,  # Производительность должна быть больше или равна требуемой
        max_temperature__gte=temperature,  # Максимальная температура не меньше требуемой
    ).order_by('head_per_stage')
    selected_main_pump = None
    required_stages = 0
    min_diff = float('inf')  # Минимальная разница между требуемым и доступным напором

    if main_pumps.exists():
        for pump in main_pumps:
            # Рассчитываем количество ступеней
            stages = required_head / pump.head_per_stage
            if stages <= pump.max_stages:
                # Рассчитываем разницу между требуемым напором и напором насоса
                diff = abs(required_head - (pump.head_per_stage * stages))
                if diff < min_diff:
                    min_diff = diff
                    selected_main_pump = pump
                    required_stages = int(stages) + 1 if not stages.is_integer() else int(stages)

    # Если насос не найден, выбираем насос с минимальной разницей
    if not selected_main_pump and main_pumps.exists():
        for pump in main_pumps:
            # Рассчитываем количество ступеней
            stages = required_head / pump.head_per_stage
            if stages > pump.max_stages:
                stages = pump.max_stages
            # Рассчитываем разницу между требуемым напором и напором насоса
            diff = abs(required_head - (pump.head_per_stage * stages))
            if diff < min_diff:
                min_diff = diff
                selected_main_pump = pump
                required_stages = int(stages)

    # Подбор подпорного насоса
    selected_booster_pump = None
    if selected_main_pump:
        main_flow_rate = selected_main_pump.flow_rate
        pipeline_diameter = pipeline_data['diameter']  # Диаметр трубопровода
        # Фильтруем подпорные насосы с производительностью, близкой к основному насосу
        booster_pumps = BoosterPump.objects.filter(
            max_temperature__gte=temperature,  # Максимальная температура не меньше требуемой
        ).order_by('flow_rate')
        min_diff_flow_rate = float('inf')  # Минимальная разница между производительностями
        if booster_pumps.exists():
            main_head = selected_main_pump.head_per_stage * required_stages
            remaining_head = required_head - main_head
            for pump in booster_pumps:
                try:
                    # Получаем данные насоса из справочника
                    pump_index = NPV_PUMP_DATA["типоразмер насоса"].index(pump.name)
                    pump_d2 = NPV_PUMP_DATA["Д_2"][pump_index]  # Диаметр Д_2 для насоса
                    # Проверяем, что диаметр Д_2 насоса близок к диаметру трубопровода
                    if any(abs(d2 - pipeline_diameter) <= 10 for d2 in pump_d2):  # Допустимая погрешность 10 мм
                        diff_flow_rate = abs(main_flow_rate - pump.flow_rate)
                        if diff_flow_rate < min_diff_flow_rate and pump.head >= remaining_head:
                            min_diff_flow_rate = diff_flow_rate
                            selected_booster_pump = pump
                except ValueError:
                    # Если насоса нет в справочнике, пропускаем его
                    print(f"Насос {pump.name} отсутствует в справочнике.")
                    continue
            # Если подходящий подпорный насос не найден, выбираем насос с минимальной разницей
            if not selected_booster_pump and booster_pumps.exists():
                for pump in booster_pumps:
                    diff_flow_rate = abs(main_flow_rate - pump.flow_rate)
                    if diff_flow_rate < min_diff_flow_rate:
                        min_diff_flow_rate = diff_flow_rate
                        selected_booster_pump = pump
    # Исходные данные
    P_pg = round(most_viscous_product['density_at_temperature'], 1)  # Плотность топлива, округлено до десятых
    hourly_rate = calculate_hourly_rate(updated_product_data)  # Расход (м³/ч)
    initial_diameter = pipeline_data['diameter']  # Начальный диаметр (мм)
    current_diameter = initial_diameter

    # Константы
    m = 3  # Константа
    g = 9.81  # Ускорение свободного падения

    # Итерационный процесс подбора диаметра
    while True:
        # Подбор основного насоса
        if selected_main_pump:
            pump_data_list = NM_PUMP_DATA.get(selected_main_pump.name, [])
            closest_pump_data = None
            min_diff = float('inf')

            for rotor_value in [1, 0.7, 1.25]:
                for pump_data in pump_data_list:
                    if pump_data.get("Rotor") == rotor_value:
                        D2 = pump_data.get("D₂", 0)
                        if D2 is not None and D2 <= current_diameter:  # Выбираем только меньшие или равные диаметры
                            diff = abs(current_diameter - D2)  # Вычисляем разницу между текущим диаметром и D₂
                            if diff < min_diff:  # Если разница меньше минимальной
                                min_diff = diff
                                closest_pump_data = pump_data

            if closest_pump_data:
                H_0 = closest_pump_data.get("H_n", 0)
                a_0 = closest_pump_data.get("a", 0)
                Q = hourly_rate
                # Расчет h_mn
                h_mn_main1 = H_0 - (a_0 * 10 ** -6 * Q ** 2)
                context['h_mn_main1'] = h_mn_main1  # Обновляем значение h_mn_main1 в контексте

        # Подбор подпорного насоса
        if selected_booster_pump:
            try:
                pump_index = NPV_PUMP_DATA["типоразмер насоса"].index(selected_booster_pump.name)
                if "Д_2" in NPV_PUMP_DATA and pump_index < len(NPV_PUMP_DATA["Д_2"]):
                    pump_d2s = NPV_PUMP_DATA["Д_2"][pump_index]
                else:
                    context['h2_booster'] = None
                    continue

                closest_d2_index = -1
                max_smaller_d2 = -float('inf')  # Находим максимальный диаметр, который меньше current_diameter
                min_d2 = min(pump_d2s)  # Минимальное значение d2 в справочнике
                for i, d2 in enumerate(pump_d2s):
                    if d2 <= current_diameter:  # Выбираем только меньшие или равные диаметры
                        if d2 > max_smaller_d2:
                            max_smaller_d2 = d2
                            closest_d2_index = i

                # Если нет подходящих d2, выбираем минимальное значение
                if closest_d2_index == -1:
                    closest_d2_index = pump_d2s.index(min_d2)


                H_0 = NPV_PUMP_DATA["Н_0,м"][pump_index][closest_d2_index]
                b_0 = NPV_PUMP_DATA["10^6*b ,ч/м^2"][pump_index][closest_d2_index]
                Q = hourly_rate
                # Расчет h2
                h2_booster = H_0 - (b_0 * 10 ** -6 * Q ** 2)
                context['h2_booster'] = h2_booster  # Обновляем значение h2_booster в контексте
            except ValueError:
                context['h2_booster'] = None

        # Расчет P
        h_mn_main1 = context.get('h_mn_main1', 0)  # Берем значение h_mn_main1 из контекста
        h2_booster = context.get('h2_booster', 0)  # Берем значение h2_booster из контекста


        P = P_pg * g * (m * h_mn_main1 + h2_booster)
        P_in_millions = round(P / 1e6, 2)



        # Проверка условия P < 6.4 МПа
        if P_in_millions < 6.4:
            context['condition_message'] = "Условие P < 6.4 МПа выполняется."
            break  # Условие выполнено, выходим из цикла
        else:
            # Уменьшаем диаметр на 10 мм и повторяем расчет
            current_diameter -= 10
            context[
                'condition_message'] = f"Условие P < 6.4 МПа не выполняется. Уменьшаем диаметр до {current_diameter} мм."



        # Добавляем расчет h_st = 3 * h_mn
    h_st = 3 * h_mn_main1  # Вычисляем h_st
    context['h_st'] = h_st  # Добавляем результат в контекст
    context['h_st_formula'] = {
        "formula": f"h_st = 3 * h_mn = 3 * {h_mn_main1}",
        "description": "Формула для h_st, где:\n"
                       f"- h_mn = высота основного бака ({h_mn_main1})",
    }

    # Добавляем результат и формулу в контекст
    context['P'] = P_in_millions
    context['P_formula'] = {
        "formula": f"P = {P_pg} * {g} * ({m} * {(h_mn_main1)} + {context.get('h2_booster', 0)})",
        "description": "Формула для P, где:\n"
                       f"- P_pg = плотность самого вязкого топлива ({P_pg})\n"
                       f"- g = ускорение свободного падения ({g})\n"
                       f"- m = константа ({m})\n"
                       f"- h_mn_main = высота основного бака ({context.get('h_mn_main1', 0)})\n"
                       f"- h2_booster = высота бустера ({context.get('h2_booster', 0)})",
    }

      # Добавляем расчет h_st = friction_head_loss * h2_booster/h_st
    n = (total_head_loss - 2* h2_booster)/ h_st  # Вычисляем h_st
    context['n'] = n  # Добавляем результат в контекст
    context['n_Formula'] = {
        "formula": f"h_st = friction_head_loss * h2_booster/h_st = {friction_head_loss} * {h2_booster}",
    }
        # Преобразуем объект MainPump в словарь
    if selected_main_pump:
        selected_main_pump_dict = selected_main_pump.to_dict()
    else:
        selected_main_pump_dict = None


    # Формируем отчет для сохранения в сессии
    report_data = {
        'pipeline_data': pipeline_data,
        'product_data': updated_product_data,
        'total_quantity': total_quantity,
        'hourly_rate': hourly_rate,
        'flow_rate': q_flow_rate,
        'total_head_loss':total_head_loss,
        'velocity': v_rounded,
        'reynolds_number': Re,
        'relative_roughness': epsilon,
        'transition_reynolds_i': Re_i,
        'transition_reynolds_ii': Re_ii,
        'hydraulic_resistance': lambda_rounded,
        'friction_head_loss': friction_head_loss,
        'most_viscous_product': most_viscous_product,
        'required_stages': required_stages,
        'selected_booster_pump': selected_booster_pump.to_dict() if selected_booster_pump else None,
        'min_diff': min_diff,
        'min_diff_booster': min_diff_flow_rate,
        'h_mn_main1': h_mn_main1,
        'P': P_in_millions,
        'h_st': h_st,
        'n': n,
        'selected_main_pump': selected_main_pump_dict,
    }

    if 'reports' not in request.session:
        request.session['reports'] = []
    request.session['reports'].append(report_data)
    request.session.modified = True  # Обязательно для сохранения изменений в сессии
    report_id = save_report_to_file(report_data, request.user.id)

    # Обновляем контекст для отображения
    context.update(report_data)

    return render(request, 'trans/result.html', context)



def archive(request):
    report_dir = f"reports/user_{request.user.id}"
    reports = []
    if os.path.exists(report_dir):
        for report_file in os.listdir(report_dir):
            with open(os.path.join(report_dir, report_file), 'r') as f:
                report_data = json.load(f)
                report_id = int(report_file.split('_')[1].split('.')[0])  # Extract the id from the filename
                report_data['id'] = report_id  # Ensure the id is set
                reports.append(report_data)
    return render(request, 'trans/archive.html', {'reports': reports})

def save_report_to_file(report_data, user_id):
    report_dir = f"reports/user_{user_id}"
    os.makedirs(report_dir, exist_ok=True)
    report_id = len(os.listdir(report_dir)) + 1  # Generate a unique ID
    report_data['id'] = report_id  # Add the ID to the report data
    report_path = os.path.join(report_dir, f"report_{report_id}.json")
    with open(report_path, 'w') as f:
        json.dump(report_data, f)
    return report_id


def report_detail(request, report_id):
    report_dir = f"reports/user_{request.user.id}"
    report_path = os.path.join(report_dir, f"report_{report_id}.json")

    if os.path.exists(report_path):
        with open(report_path, 'r') as f:
            report_data = json.load(f)
        return render(request, 'trans/report_detail.html', {'report': report_data, 'report_id': report_id})
    else:
        return HttpResponse("Отчет не найден", status=404)




def delete_report(request, report_id):
    # Construct the path to the report file
    report_dir = f"reports/user_{request.user.id}"
    report_path = os.path.join(report_dir, f"report_{report_id}.json")

    if os.path.exists(report_path):
        os.remove(report_path)

        if os.path.exists(report_dir) and not os.listdir(report_dir):
            os.rmdir(report_dir)  # Delete the directory if it's empty

    return redirect('trans:archive')





@login_required
def download_report(request, report_id):
    import os
    # Получаем данные из сессии
    reports = request.session.get('reports', [])
    if report_id >= len(reports):
        return HttpResponse("Отчет не найден", status=404)
    report = reports[report_id]




    # Создаем Excel-документ
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"

    # Заголовок отчета
    ws['A1'] = "Отчет по трубопроводу"
    ws['A1'].font = Font(bold=True, size=14)  # Используем объект Font

    # Основные данные о трубопроводе
    ws['A3'] = "Основные данные о трубопроводе:"
    ws['A3'].font = Font(bold=True)  # Используем объект Font
    ws['A4'] = f"Диаметр: {report['pipeline_data']['diameter']} мм"
    ws['A5'] = f"Длина: {report['pipeline_data']['length']} км"
    ws['A6'] = f"Температура: {report['pipeline_data']['temperature']} K"
    ws['A7'] = f"Общее количество: {report['total_quantity']} млн. т/год"

    # Данные по продуктам
    ws['A9'] = "Данные по продуктам:"
    ws['A9'].font = Font(bold=True)  # Используем объект Font
    ws['A10'] = "Продукт"
    ws['B10'] = "Плотность при 293K (кг/м³)"
    ws['C10'] = f"Плотность при {report['pipeline_data']['temperature']}K (кг/м³)"
    ws['D10'] = "Вязкость при 273K (мм²/с)"
    ws['E10'] = "Вязкость при 293K (мм²/с)"
    ws['F10'] = "Процент (%)"
    ws['G10'] = "Количество (млн. т/год)"

    row = 11
    for product in report['product_data']:
        ws[f'A{row}'] = product['name']
        ws[f'B{row}'] = product['density293']
        ws[f'C{row}'] = product.get('density_at_temperature', 'N/A')
        ws[f'D{row}'] = product['viscosity273']
        ws[f'E{row}'] = product['viscosity293']
        ws[f'F{row}'] = product['percentage']
        ws[f'G{row}'] = product['quantity']
        row += 1

    # Результаты расчетов
    ws[f'A{row + 2}'] = "Результаты расчетов:"
    ws[f'A{row + 2}'].font = Font(bold=True)  # Используем объект Font
    ws[f'A{row + 3}'] = "Число Рейнольдса"
    ws[f'B{row + 3}'] = report['reynolds_number']
    ws[f'A{row + 4}'] = "Относительная шероховатость"
    ws[f'B{row + 4}'] = report['relative_roughness']
    ws[f'A{row + 5}'] = "Коэффициент гидравлического сопротивления"
    ws[f'B{row + 5}'] = report['hydraulic_resistance']
    ws[f'A{row + 6}'] = "Потеря напора на трение"
    ws[f'B{row + 6}'] = report['friction_head_loss']
    ws[f'A{row + 7}'] = "Полные потери напора"
    ws[f'B{row + 7}'] = report['total_head_loss']
    ws[f'A{row + 8}'] = "P"
    ws[f'B{row + 8}'] = report['P']

    # Данные по насосам
    ws[f'A{row + 10}'] = "Данные по насосам:"
    ws[f'A{row + 10}'].font = Font(bold=True)  # Используем объект Font
    if report['selected_main_pump']:
        ws[f'A{row + 11}'] = "Основной насос:"
        ws[f'A{row + 11}'].font = Font(bold=True)  # Используем объект Font
        ws[f'A{row + 12}'] = "Название"
        ws[f'B{row + 12}'] = report['selected_main_pump']['name']
        ws[f'A{row + 13}'] = "Производительность"
        ws[f'B{row + 13}'] = report['selected_main_pump']['flow_rate']
        ws[f'A{row + 14}'] = "Напор на ступень"
        ws[f'B{row + 14}'] = report['selected_main_pump']['head_per_stage']
        ws[f'A{row + 15}'] = "Количество ступеней"
        ws[f'B{row + 15}'] = report['required_stages']
        ws[f'A{row + 16}'] = "Мощность на ступень"
        ws[f'B{row + 16}'] = report['selected_main_pump']['power_per_stage']
        ws[f'A{row + 17}'] = "КПД"
        ws[f'B{row + 17}'] = report['selected_main_pump']['efficiency']
    else:
        ws[f'A{row + 11}'] = "Основной насос не найден."

    if report['selected_booster_pump']:
        ws[f'A{row + 19}'] = "Подпорный насос:"
        ws[f'A{row + 19}'].font = Font(bold=True)  # Используем объект Font
        ws[f'A{row + 20}'] = "Название"
        ws[f'B{row + 20}'] = report['selected_booster_pump']['name']
        ws[f'A{row + 21}'] = "Производительность"
        ws[f'B{row + 21}'] = report['selected_booster_pump']['flow_rate']
        ws[f'A{row + 22}'] = "Напор"
        ws[f'B{row + 22}'] = report['selected_booster_pump']['head']
        ws[f'A{row + 23}'] = "Мощность"
        ws[f'B{row + 23}'] = report['selected_booster_pump']['power']
        ws[f'A{row + 24}'] = "КПД"
        ws[f'B{row + 24}'] = report['selected_booster_pump']['efficiency']
    else:
        ws[f'A{row + 19}'] = "Подпорный насос не найден."

    # Сохраняем Excel-файл
    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Возвращаем Excel как ответ
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="report_{report_id}.xlsx"'
    return response

    # архив представлений





@login_required
def calculator(request):
    products = Product.objects.all()
    error = None
    print(f"User is authenticated: {request.user.is_authenticated}")
    if not request.user.is_authenticated:
        return redirect('trans:login')  # Перенаправляем на страницу входа

    if request.method == 'POST':
        try:
            # Получаем общее количество из формы
            total_quantity = float(request.POST.get('quantity', 0))

            # Создаем объект Pipeline с данными из формы
            pipeline = Pipeline.objects.create(
                user=request.user,  # Добавляем текущего пользователя
                Diameter=float(request.POST.get('diameter')),
                Length=float(request.POST.get('length')),
                ElevationDifference=float(request.POST.get('elevation_difference')),
                ResidualHead=float(request.POST.get('residual_head')),
                Temperature=float(request.POST.get('temperature')),
            )

            # Обрабатываем продукты
            product_ids = request.POST.getlist('products')
            percentages = request.POST.getlist('product-percentage')
            quantities = request.POST.getlist('product-quantity')

            # Список для хранения данных о продуктах
            product_data = []

            for product_id, percentage, quantity in zip(product_ids, percentages, quantities):
                product = Product.objects.get(ProductID=product_id)
                # Добавляем данные о продукте в список
                product_data.append({
                    'name': product.Name,
                    'density293': product.Density293,
                    'viscosity273': product.Viscosity273,
                    'viscosity293': product.Viscosity293,
                    'percentage': float(percentage),
                    'quantity': float(quantity),
                })

            # Сохраняем данные в сессии для передачи на result.html
            request.session['pipeline_data'] = {
                'diameter': pipeline.Diameter,
                'length': pipeline.Length,
                'elevation_difference': pipeline.ElevationDifference,
                'residual_head': pipeline.ResidualHead,
                'temperature': pipeline.Temperature,
            }
            request.session['product_data'] = product_data
            request.session['total_quantity'] = total_quantity

            # Перенаправляем на страницу результата
            return redirect('trans:result')

        except Exception as e:
            error = f"Ошибка: {str(e)}"

    return render(request, 'trans/calculator.html', {
        'products': products,
        'error': error,
    })




def home(request):
    return render(request, 'trans/home.html')







def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('trans:dashboard')  # Используем namespace 'trans'
    else:
        form = RegisterForm()
    return render(request, 'trans/register.html', {'form': form})

from django.utils import timezone
from django.contrib.auth import authenticate, login

def user_login(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            request.session['login_time'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
            if user is not None:
                login(request, user)
                # Перенаправляем в зависимости от роли
                if user.role == 'engineer':
                    return redirect('trans:engineer_dashboard')  # Используем namespace 'trans'
                elif user.role == 'admin':
                    return redirect('trans:admin_dashboard')  # Используем namespace 'trans'
    else:
        form = LoginForm()
    return render(request, 'trans/login.html', {'form': form})



def custom_logout(request):
    if request.method == 'POST' or request.method == 'GET':  # Allow both GET and POST
        logout(request)
        return redirect('trans:home')  # Redirect to the home page after logout
    else:
        return HttpResponseNotAllowed(['GET', 'POST'])  # Return 405 for unsupported methods


@login_required
def dashboard(request):
    user = request.user
    pipelines = user.pipelines.all()
    if user.role == 'engineer':
        return render(request, 'trans/engineer_dashboard.html', {'user': user, 'pipelines': pipelines})
    elif user.role == 'admin':
        return render(request, 'trans/admin_dashboard.html', {'user': user, 'pipelines': pipelines})




from django.utils import timezone

from django.utils import timezone

def session_info(request):
    if request.user.is_authenticated:
        return {
            'session_info': {
                'role': request.user.get_role_display(),
                'full_name': f"{request.user.surname} {request.user.name} {request.user.patronymic}",
                'login_time': request.session.get('login_time', 'N/A'),
            }
        }
    return {}




